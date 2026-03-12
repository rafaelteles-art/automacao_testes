"""
Simulate filling PRÉ-ESCALA section (fixed encoding + FB debug)
"""
import re, sys, io, requests
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FB_TOKEN = "EAAWDHozjODgBQ0b4ZAZBOZBzGhqi9ZCX0bj8DbmAPnsBfYbEMMZCqMeBMCmLjB2dpzxHvzZC6UQGApi9frZAWyQHPmHZB1hFJa2q3nTNaaDtwHSxqJB5Veeo1CpE9gTYAD3vpJf9vRNNj62z2ebVJ6tD0mKbIzh9DXZCbrnjOHhiAkrcffsEwKcZAuHchAMZBRgi1BjmUIjP2IhfH7O"
RT_TOKEN = "wB7qY69R0KVU9tl4TBaQ"
EXCEL_FILE = r"C:\Preencher planilha\FB - LOTTO V7.xlsx"
DATE_START = "2026-03-02"
DATE_END = "2026-03-06"

accounts = [
    "987248712293933", "1583963688753767", "6481076258591934",
    "1230367020974448", "542987171356461", "1300700817490669",
    "854290605799299", "982787082695569", "875839750407156",
    "550719079377709", "1277255524217242"
]

# Read Excel
wb = load_workbook(EXCEL_FILE, read_only=True)
ws = wb["032026"]
rows = []
for row_idx in range(4, ws.max_row + 1):
    creative = ws.cell(row=row_idx, column=15).value
    status = ws.cell(row=row_idx, column=22).value
    if creative and str(creative).strip():
        rows.append({'row': row_idx, 'creative': str(creative).strip(), 'status': str(status).strip().upper() if status else ''})
wb.close()
print(f"Excel: {len(rows)} rows in Col O")

# USD/BRL
try:
    usd_to_brl = float(requests.get("https://economia.awesomeapi.com.br/last/USD-BRL").json()["USDBRL"]["ask"])
except:
    usd_to_brl = 5.0
print(f"Dolar: R$ {usd_to_brl:.4f}")

# Facebook
print("\nFetching Facebook campaigns...")
def extract_ad(name):
    if not name: return ""
    m = re.search(r'(?:ABO|CBO)\s+\S+\s*-\s*(.+)$', name, re.IGNORECASE)
    if m: return m.group(1).strip()
    parts = name.rsplit(' - ', 1)
    return parts[1].strip() if len(parts) == 2 else ""

ad_catalog = {}
finance = {}

for acc in accounts:
    # Test token first for first account only
    url = f"https://graph.facebook.com/v19.0/act_{acc}/campaigns"
    p = {"access_token": FB_TOKEN, "fields": "id,name", "limit": 500}
    while url:
        r = requests.get(url, params=p, timeout=30)
        d = r.json()
        if 'error' in d:
            print(f"  FB Error on {acc}: {d['error'].get('message', 'unknown')[:80]}")
            break
        for c in d.get("data", []):
            ex = extract_ad(c.get("name", ""))
            if ex:
                k = ex.strip().lower()
                k_clean = k.strip('[]')
                if k_clean not in ad_catalog:
                    ad_catalog[k_clean] = {"id": c["id"], "name": c["name"]}
                if k != k_clean and k not in ad_catalog:
                    ad_catalog[k] = {"id": c["id"], "name": c["name"]}
        url = d.get("paging", {}).get("next")
        p = {}
    
    # Insights
    url = f"https://graph.facebook.com/v19.0/act_{acc}/insights"
    p = {
        'access_token': FB_TOKEN,
        'fields': 'campaign_id,spend',
        'level': 'campaign',
        'time_range': f'{{"since":"{DATE_START}","until":"{DATE_END}"}}',
        'limit': 100
    }
    while url:
        r = requests.get(url, params=p, timeout=30)
        d = r.json()
        if 'error' in d:
            break
        for row in d.get('data', []):
            cid = row.get("campaign_id")
            if cid:
                existing = finance.get(cid, {"spend": 0.0})
                existing["spend"] += float(row.get("spend", 0.0))
                finance[cid] = existing
        url = d.get("paging", {}).get("next")
        p = {}

print(f"  {len(ad_catalog)} campaign keys, {len(finance)} with insights")

# RedTrack
print("Fetching RedTrack...")
rt_map = {}
page = 1
while True:
    r = requests.get('https://api.redtrack.io/report', params={
        'api_key': RT_TOKEN, 'date_from': DATE_START, 'date_to': DATE_END,
        'group': 'rt_ad', 'limit': 1000, 'page': page
    }, timeout=30)
    if r.status_code != 200: break
    data = r.json()
    for row in data:
        rt_ad = str(row.get('rt_ad', '')).strip().lower()
        if rt_ad:
            rt_map[rt_ad] = {
                'vendas': float(row.get('convtype2', 0)),
                'roas': float(row.get('roas', 0)),
            }
    if len(data) < 1000: break
    page += 1
print(f"  {len(rt_map)} rt_ad entries\n")

# Simulate
print("=" * 115)
print(f"{'Row':<5} | {'Creative (O)':<18} | {'Status (V)':<12} | {'Gasto R$ (R)':<14} | {'Vendas (S)':<10} | {'ROAS (T)':<10} | {'CPA R$ (U)':<14}")
print("-" * 115)

for r in rows:
    creative = r['creative']
    status = r['status']
    search = creative.lower()
    
    # Collect ALL matching campaign IDs
    matched_ids = []
    
    if search in ad_catalog:
        matched_ids.append(ad_catalog[search]['id'])
    
    # Bracketed
    bk = f"[{search}]"
    if bk in ad_catalog and ad_catalog[bk]['id'] not in matched_ids:
        matched_ids.append(ad_catalog[bk]['id'])
    
    # Variations: LT1011 matches LT1011.2, LT1011.4, etc.
    for key, info in ad_catalog.items():
        clean = key.strip('[]')
        if clean.startswith(search + '.') or clean.startswith(search + ' '):
            if info['id'] not in matched_ids:
                matched_ids.append(info['id'])
    
    total_spend_usd = sum(finance.get(cid, {}).get('spend', 0) for cid in matched_ids)
    spend_brl = total_spend_usd * usd_to_brl
    
    rt = rt_map.get(search, None)
    vendas = rt['vendas'] if rt else 0
    roas = rt['roas'] if rt else 0
    cpa = spend_brl / vendas if vendas > 0 else (spend_brl if spend_brl > 0 else 0)
    
    icon = "[OK]" if "TESTE" in status else "[SKIP]"
    print(f"{icon} {r['row']:<3} | {creative:<18} | {status:<12} | R$ {spend_brl:<10.2f} | {vendas:<10.0f} | {roas:<10.4f} | R$ {cpa:<10.2f}")

print("=" * 115)
