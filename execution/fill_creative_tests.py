#!/usr/bin/env python3
"""
fill_creative_tests.py
Fills Column A (TC group label) and performance metrics (Hook, Body75, CPM, CTR, CPC, Gasto)
of the '032026' sheet conditionally based on User Column M (Status).

Logic:
1. Fetches USD->BRL quote from AwesomeAPI.
2. Fetches ALL campaigns directly via /campaigns to parse out names and TCs.
3. Reads each Excel test row, parses its specific 'Início' date (Col C / Col P).
4. Fetches targeted /insights and /report (RedTrack) dynamically for that exact date range.
5. Col A Logic: Fills Col A if blank and not already merged.
6. Metrics Logic: Fills financial and performance columns ONLY if Col M contains 'TESTE'.
"""

import re
import sys
import os
import requests
import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

EXCEL_FILE = r"C:\Preencher planilha\FB - LOTTO V7.xlsx"
MAIN_SHEET = "032026"
DATA_START_ROW = 4

DEFAULT_FB_TOKEN = "EAAWDHozjODgBQ6y16hDYtG5psMsSYZBv3S6O8w3lxeihOtCryNgLFwQXGVxwbJuSks9BJQZBsWY2Iw5ZAuV9sCuTmBusecWWrkxb5glHaKLE9DSvhe6edmfT0b96sVWGZAsBVwZCMAF9cKsPJQ4qZBEGJWFoOZA9iDAPZBFT5GNUMRIIJEjj0hik5d1nTZAvWcjNMwMurcmcWgOfXOWtaZBaM3nYZB2EbdqllBH3mAjCxSGkAgocfYvJZBZCkZCdfmqUcY5LekdzG3w6doGIB0IytWhbY9JZBTEepCY5DgS3ZB8OXQZDZD"

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from facebook_redtrack_importer_v2 import FacebookAdsAPI

def parse_campaign_name(campaign_name: str):
    if not campaign_name: return None
    tc_match = re.search(r'TC(\d+)', campaign_name, re.IGNORECASE)
    if not tc_match: return None
    tc_label = f"TC{tc_match.group(1)}"
    bracket_match = re.search(r'\]([^[]*?)TC\d+', campaign_name, re.IGNORECASE)
    account_name = bracket_match.group(1).strip() if bracket_match else ""
    return tc_label, account_name

def extract_ad_name_from_campaign(campaign_name: str) -> str:
    if not campaign_name: return ""
    match = re.search(r'(?:ABO|CBO)\s+\S+\s*-\s*(.+)$', campaign_name, re.IGNORECASE)
    if match: return match.group(1).strip()
    parts = campaign_name.rsplit(' - ', 1)
    if len(parts) == 2: return parts[1].strip()
    return ""

def build_col_a_label(campaign_name: str) -> str:
    result = parse_campaign_name(campaign_name)
    if not result: return ""
    return f"{result[0]} {result[1]}".strip()

def parse_excel_date(cell_value, default_date: str) -> str:
    """Safely parse Excel datetime objects or strings to YYYY-MM-DD"""
    if not cell_value:
        return default_date
    if isinstance(cell_value, datetime.datetime):
        return cell_value.strftime("%Y-%m-%d")
    if isinstance(cell_value, str):
        # Handle 'dd/mm/yyyy' typical in BR Excel
        match = re.search(r'(\d{2})/(\d{2})/(\d{4})', cell_value)
        if match:
            return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
        # Handle 'dd/mm/yy'
        match_short = re.search(r'(\d{2})/(\d{2})/(\d{2})', cell_value)
        if match_short:
            return f"20{match_short.group(3)}-{match_short.group(2)}-{match_short.group(1)}"
    return default_date

def fetch_fb_insights_for_campaign(c_id, since, until, fb_token):
    url = f"https://graph.facebook.com/v19.0/{c_id}/insights"
    params = {
        'access_token': fb_token,
        'fields': 'campaign_id,impressions,cpc,cpm,ctr,spend,actions,video_p75_watched_actions',
        'level': 'campaign',
        'time_range': f'{{"since":"{since}","until":"{until}"}}'
    }
    import time
    for _ in range(3):
        r = requests.get(url, params=params, timeout=30)
        if r.status_code == 200:
            data = r.json()
            if data.get('data'):
                row = data['data'][0]
                imps = float(row.get("impressions", 0))
                # Hook Rate
                actions = row.get('actions', [])
                video_3s_views = 0
                for act in actions:
                    if act.get('action_type') == 'video_view':
                        video_3s_views = float(act.get('value', 0))
                        break
                # Hold Rate
                video_p75_actions = row.get('video_p75_watched_actions', [])
                p75 = float(video_p75_actions[0].get('value', 0)) if video_p75_actions else 0
                
                return {
                    "spend": float(row.get("spend", 0.0)),
                    "cpm": float(row.get("cpm", 0.0)),
                    "cpc": float(row.get("cpc", 0.0)),
                    "ctr": (float(row.get("ctr", 0.0)) / 100),
                    "hook_rate": (video_3s_views / imps) if imps > 0 else 0,
                    "body_rate": (p75 / imps) if imps > 0 else 0
                }
            break
        elif r.status_code == 429 or r.status_code >= 500:
            time.sleep(2)
            continue
        else:
            break

    return {
        "spend": 0.0, "cpm": 0.0, "cpc": 0.0,
        "ctr": 0.0, "hook_rate": 0.0, "body_rate": 0.0
    }

def fetch_rt_for_ad(ad_name_lower, since, until, rt_token):
    vendas = 0.0
    cost = 0.0
    roas = 0.0
    if not rt_token or not ad_name_lower:
        return {"vendas": 0, "cost": 0.0, "roas": 0.0}
    
    import time
    page = 1
    while page <= 5: # Limit pagination for row-by-row fetching to prevent lockups
        success = False
        rt_data = []
        for _ in range(3):
            r = requests.get('https://api.redtrack.io/report', params={
                'api_key': rt_token,
                'date_from': since,
                'date_to': until,
                'group': 'rt_ad',
                'limit': 1000,
                'page': page
            }, timeout=30)
            
            if r.status_code == 200:
                rt_data = r.json()
                success = True
                break
            elif r.status_code == 429 or r.status_code >= 500:
                time.sleep(2)
                continue
            else:
                break
                
        if not success or not rt_data: break
        
        for r_row in rt_data:
            rt_ad = str(r_row.get('rt_ad', '')).strip().lower()
            if not rt_ad: continue
            
            # Match condition logic - EXACT match first, then prefix extraction
            is_match = False
            if rt_ad == ad_name_lower:
                is_match = True
            elif rt_ad == ad_name_lower.split(" - ")[0].split(" ")[0]:
                is_match = True
                
            if is_match:
                vendas += float(r_row.get('convtype2', 0))
                cost += float(r_row.get('cost', 0))
                roas_val = float(r_row.get('roas', 0))
                if roas_val != 0: roas = roas_val
        
        if len(rt_data) < 1000: break
        page += 1

    return {"vendas": vendas, "cost": cost, "roas": roas}

def fill_creative_tests(
    account_ids: list, 
    date_start: str, 
    date_end: str, 
    fb_token: str = None,
    redtrack_token: str = None,
    fb_api_instance=None, 
    progress_callback=None
):
    token = fb_token or DEFAULT_FB_TOKEN

    # 1. Fetch Dolar Quote
    usd_to_brl = 5.0
    try:
        r_quote = requests.get("https://economia.awesomeapi.com.br/last/USD-BRL")
        usd_to_brl = float(r_quote.json()["USDBRL"]["ask"])
        if progress_callback: progress_callback(f"Cotação do Dólar obtida: R$ {usd_to_brl:.4f}")
    except Exception as e:
        if progress_callback: progress_callback(f"⚠️ Erro ao obter dólar. Usando R$ 5.00.")

    # 2. Fetch all campaigns to build catalog
    if progress_callback: progress_callback("Construindo catálogo raiz de Campanhas...")

    all_campaigns = []
    for account_id in account_ids:
        acc_raw = account_id.replace("act_", "")
        url = f"https://graph.facebook.com/v19.0/act_{acc_raw}/campaigns"
        params = {"access_token": token, "fields": "id,name", "limit": 500}
        while url:
            r = requests.get(url, params=params, timeout=30)
            if r.status_code != 200: break
            data = r.json()
            page_data = data.get("data", [])
            if not page_data: break
            all_campaigns.extend(page_data)
            paging = data.get("paging", {})
            url = paging.get("next")
            params = {}

    ad_to_campaign = {}
    for camp in all_campaigns:
        c_name = camp.get("name", "")
        c_id = camp.get("id")
        if c_name:
            extracted = extract_ad_name_from_campaign(c_name)
            if extracted:
                key = extracted.strip().lower()
                if key not in ad_to_campaign: ad_to_campaign[key] = {"id": c_id, "name": c_name}

    if not ad_to_campaign:
        raise RuntimeError("Nenhuma campanha válida encontrada para as contas selecionadas.")

    # 3. Read Excel and Process Row-by-Row
    if progress_callback:
        progress_callback("Varrendo planilha e extraindo dados específicos por linha (Dynamic Dates)...")

    try:
        wb = load_workbook(EXCEL_FILE)
    except PermissionError:
        raise PermissionError(f"O Arquivo Excel está aberto. Feche a planilha e tente novamente.")
    
    ws = wb[MAIN_SHEET]

    filled_a = 0
    filled_metrics = 0
    skipped_rows = 0
    not_found = []

    # Process TESTES Completos Section
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        ad_name_cell = ws.cell(row=row_idx, column=2)
        ad_name_value = ad_name_cell.value
        if not ad_name_value or str(ad_name_value).strip() == "":
            continue

        search_term = str(ad_name_value).strip().lower()
        matched_info = None
        
        if search_term in ad_to_campaign:
            matched_info = ad_to_campaign[search_term]
        else:
            for key, info in ad_to_campaign.items():
                if search_term in key or key in search_term:
                    matched_info = info
                    break

        if not matched_info:
            not_found.append(str(ad_name_value))
            continue
            
        c_name = matched_info["name"]
        c_id = matched_info["id"]

        # Col A Logic
        cell_a = ws.cell(row=row_idx, column=1)
        if type(cell_a).__name__ != 'MergedCell':
            if not cell_a.value or str(cell_a.value).strip() == "":
                label = build_col_a_label(c_name)
                if label:
                    cell_a.value = label
                    cell_a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell_a.font = Font(bold=True)
                    filled_a += 1

        # Check Col M Status
        status_cell = ws.cell(row=row_idx, column=13)
        current_status = str(status_cell.value).strip().upper() if status_cell.value else ""
        
        if "TESTE" in current_status:
            # Get specific start date from Col C (column 3)
            date_col_c = ws.cell(row=row_idx, column=3).value
            row_date_start = parse_excel_date(date_col_c, date_start)
            
            # Fetch FB data dynamically for this row
            fin = fetch_fb_insights_for_campaign(c_id, row_date_start, date_end, token)
            
            cpc_brl = fin["cpc"] * usd_to_brl
            cpm_brl = fin["cpm"] * usd_to_brl
            spend_brl = fin["spend"] * usd_to_brl
            
            # Write metrics...
            c5 = ws.cell(row=row_idx, column=5, value=fin["hook_rate"])
            c5.number_format = '0.00%'
            c6 = ws.cell(row=row_idx, column=6, value=fin["body_rate"])
            c6.number_format = '0.00%'
            c7 = ws.cell(row=row_idx, column=7, value=cpm_brl)
            c7.number_format = '#,##0.00'
            c8 = ws.cell(row=row_idx, column=8, value=fin["ctr"])
            c8.number_format = '0.00%'
            c9 = ws.cell(row=row_idx, column=9, value=cpc_brl)
            c9.number_format = '#,##0.00'
            c10 = ws.cell(row=row_idx, column=10, value=spend_brl)
            c10.number_format = '#,##0.00'
            
            # Fetch RedTrack dynamically for this row
            rt = fetch_rt_for_ad(search_term, row_date_start, date_end, redtrack_token)
            vendas = rt["vendas"]
            
            ws.cell(row=row_idx, column=11, value=vendas)
            
            cpa = 0
            if vendas > 0:
                cpa = spend_brl / vendas
            else:
                cpa = 0 # Forced fallback to 0
                
            c12 = ws.cell(row=row_idx, column=12, value=cpa)
            c12.number_format = '#,##0.00'

            filled_metrics += 1
        else:
            skipped_rows += 1

    # Re-merge Col A
    merge_groups = 0
    row = DATA_START_ROW
    while row <= ws.max_row:
        cell_val = ws.cell(row=row, column=1).value
        if not cell_val:
            row += 1
            continue
        end_row = row
        while end_row + 1 <= ws.max_row and ws.cell(row=end_row + 1, column=1).value == cell_val:
            end_row += 1
        if end_row > row:
            ws.merge_cells(start_row=row, start_column=1, end_row=end_row, end_column=1)
            merged_cell = ws.cell(row=row, column=1)
            merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            merged_cell.font = Font(bold=True)
            merge_groups += 1
        row = end_row + 1

    # Process PRÉ-ESCALA Section
    filled_pre_escala = 0
    skipped_pre_escala = 0
    if progress_callback:
        progress_callback("Preenchendo seção PRÉ-ESCALA com datas dinâmicas...")

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        creative_cell = ws.cell(row=row_idx, column=15)  # Col O
        creative_val = creative_cell.value
        if not creative_val or str(creative_val).strip() == "":
            continue

        status_pe = ws.cell(row=row_idx, column=22)  # Col V
        status_pe_str = str(status_pe.value).strip().upper() if status_pe.value else ""
        if "TESTE" not in status_pe_str:
            skipped_pre_escala += 1
            continue

        search_pe = str(creative_val).strip().lower()
        
        # Determine Row Date - Col P uses column 16
        date_col_p = ws.cell(row=row_idx, column=16).value
        # For Pré-Escala, parse_excel_date reads from Col P
        row_date_pe = parse_excel_date(date_col_p, date_start)
        
        # RedTrack Fetch
        rt = fetch_rt_for_ad(search_pe, row_date_pe, date_end, redtrack_token)
        
        cost_brl = rt['cost']
        vendas_pe = rt['vendas']
        roas_pe = rt['roas']
        
        cpa_pe = 0
        if vendas_pe > 0:
            cpa_pe = cost_brl / vendas_pe
        else:
            cpa_pe = 0 # forced fallback
        
        c_r = ws.cell(row=row_idx, column=18, value=cost_brl)
        c_r.number_format = '#,##0.00'
        
        ws.cell(row=row_idx, column=19, value=vendas_pe)
        
        c_t = ws.cell(row=row_idx, column=20, value=roas_pe)
        c_t.number_format = '0.00'
        
        c_u = ws.cell(row=row_idx, column=21, value=cpa_pe)
        c_u.number_format = '#,##0.00'

        filled_pre_escala += 1

    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        raise PermissionError(f"Não foi possível salvar! O arquivo '{EXCEL_FILE}' está aberto no Excel.")

    return {
        "filled_a": filled_a, 
        "filled_metrics": filled_metrics, 
        "skipped_rows": skipped_rows, 
        "not_found": not_found,
        "filled_pre_escala": filled_pre_escala,
        "skipped_pre_escala": skipped_pre_escala
    }
