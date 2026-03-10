import openpyxl

wb = openpyxl.load_workbook(r'C:\Preencher planilha\FB - LOTTO V7.xlsx', data_only=True)
ws = wb['032026']
test_ads = ['lt1017.7', 'lt1010.2', 'lt899.22', 'lt1033.2', 'lt1011.4', 'lt899.32', 'lt899.33', 'lt1034.4', 'lt1069', 'lt1070']

print('Valores atuais na Planilha (Col K e L):')
for r in range(4, ws.max_row + 1):
    ad = ws.cell(row=r, column=2).value
    if not ad: continue
    
    ad_str = str(ad).strip()
    if any(t in ad_str.lower() for t in test_ads):
        vendas = ws.cell(row=r, column=11).value
        cpa = ws.cell(row=r, column=12).value
        vendas_str = str(vendas) if vendas is not None else "N/A"
        cpa_formatted = f"R$ {cpa:.2f}" if isinstance(cpa, (int, float)) else str(cpa)
        print(f"Row {r:4d} | Anúncio: {ad_str:<35} | Vendas: {vendas_str:<5} | CPA: {cpa_formatted}")
