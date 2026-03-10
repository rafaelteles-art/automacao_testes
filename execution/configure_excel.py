#!/usr/bin/env python3
"""
Excel Configuration Script
Sets up the "Dados Brutos" sheet and adds formulas to the main sheet
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def configure_raw_data_sheet(workbook):
    """Create and format the raw data sheet"""
    
    sheet_name = "Dados Brutos"
    
    # Create sheet if it doesn't exist
    if sheet_name not in workbook.sheetnames:
        raw_sheet = workbook.create_sheet(sheet_name, 0)
    else:
        raw_sheet = workbook[sheet_name]
    
    # Define headers for Facebook data
    fb_headers = [
        "campaign_id",
        "campaign_name", 
        "adset_id",
        "adset_name",
        "ad_id",
        "ad_name",
        "impressions",
        "clicks",
        "spend",
        "cpm",
        "ctr",
        "cpc"
    ]
    
    # Define headers for RedTrack data
    rt_headers = [
        "conversion_id",
        "campaign_id",
        "revenue",
        "cpa"
    ]
    
    # Add Facebook headers
    for col_idx, header in enumerate(fb_headers, 1):
        cell = raw_sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add RedTrack headers
    for col_idx, header in enumerate(rt_headers, len(fb_headers) + 2):
        cell = raw_sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    for col_idx in range(1, len(fb_headers) + len(rt_headers) + 2):
        raw_sheet.column_dimensions[get_column_letter(col_idx)].width = 15
    
    print(f"✓ Raw data sheet '{sheet_name}' configured")
    return raw_sheet


def add_formulas_to_main_sheet(workbook):
    """Add reference formulas to the main sheet"""
    
    main_sheet_name = "032026"
    if main_sheet_name not in workbook.sheetnames:
        print(f"⚠ Main sheet '{main_sheet_name}' not found")
        return
    
    main_sheet = workbook[main_sheet_name]
    
    # Example: Add formulas in row 4 (first data row after headers)
    # These are example formulas - adjust row numbers as needed
    
    # Assuming row 3 has headers and row 4 starts data
    # Column C = CRIATIVO (from main sheet)
    
    # Example formula for CPM (column G in main, column J in raw)
    # =VLOOKUP(B4, 'Dados Brutos'!A:L, 10, FALSE)
    
    # Example formula for GASTO (column J in main, column I in raw)
    # =VLOOKUP(B4, 'Dados Brutos'!A:L, 9, FALSE)
    
    # Example formula for VENDAS (from RedTrack)
    # =SUMIF('Dados Brutos'!N:N, B4, 'Dados Brutos'!O:O)
    
    # Example formula for CPA (from RedTrack)
    # =IFERROR(VLOOKUP(B4, 'Dados Brutos'!M:P, 4, FALSE), 0)
    
    print(f"✓ Formula templates created")
    print(f"\nExample formulas to add to '{main_sheet_name}':")
    print(f"  • CPM: =VLOOKUP(B4, 'Dados Brutos'!A:L, 10, FALSE)")
    print(f"  • GASTO: =VLOOKUP(B4, 'Dados Brutos'!A:L, 9, FALSE)")
    print(f"  • VENDAS: =SUMIF('Dados Brutos'!N:N, B4, 'Dados Brutos'!O:O)")
    print(f"  • CPA: =IFERROR(VLOOKUP(B4, 'Dados Brutos'!M:P, 4, FALSE), 0)")


def main():
    """Main execution"""
    
    excel_file = "C:\Preencher planilha\FB - LOTTO V7.xlsx"
    
    print("\n" + "="*60)
    print("Excel Configuration Tool")
    print("="*60)
    
    try:
        # Load workbook
        print(f"\n1. Loading Excel file: {excel_file}")
        wb = load_workbook(excel_file)
        print(f"   ✓ Loaded sheets: {', '.join(wb.sheetnames)}")
        
        # Configure raw data sheet
        print(f"\n2. Configuring raw data sheet...")
        configure_raw_data_sheet(wb)
        
        # Add formulas to main sheet
        print(f"\n3. Preparing formulas for main sheet...")
        add_formulas_to_main_sheet(wb)
        
        # Save
        print(f"\n4. Saving Excel file...")
        wb.save(excel_file)
        print(f"   ✓ File saved successfully")
        
        print("\n" + "="*60)
        print("✓ Configuration completed!")
        print("="*60)
        print("\nNext steps:")
        print("1. Run: python3 facebook_redtrack_importer_v2.py")
        print("2. Select your Business Manager and Ad Account")
        print("3. Data will be imported to 'Dados Brutos' sheet")
        print("4. Add the example formulas to your main sheet")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        return False
    
    return True


if __name__ == "__main__":
    success = main()
    exit(0 if success else 1)
