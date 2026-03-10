#!/usr/bin/env python3
"""Read the headers of the 032026 sheet"""
import sys
from openpyxl import load_workbook

EXCEL_FILE = r"C:\Preencher planilha\FB - LOTTO V7.xlsx"
MAIN_SHEET = "032026"

wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb[MAIN_SHEET]

print("Headers na Linha 3:")
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=3, column=col).value
    col_letter = ws.cell(row=3, column=col).column_letter
    print(f"[{col_letter}] {val}")
