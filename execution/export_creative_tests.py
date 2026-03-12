#!/usr/bin/env python3
"""
export_creative_tests.py
Generates a new Excel file with Creative Tests data extracted from Campaign Names.
Returns a BytesIO object that can be downloaded via Streamlit.
"""

import re
import sys
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from facebook_redtrack_importer_v2 import FacebookAdsAPI

DEFAULT_FB_TOKEN = "EAAWDHozjODgBQ0b4ZAZBOZBzGhqi9ZCX0bj8DbmAPnsBfYbEMMZCqMeBMCmLjB2dpzxHvzZC6UQGApi9frZAWyQHPmHZB1hFJa2q3nTNaaDtwHSxqJB5Veeo1CpE9gTYAD3vpJf9vRNNj62z2ebVJ6tD0mKbIzh9DXZCbrnjOHhiAkrcffsEwKcZAuHchAMZBRgi1BjmUIjP2IhfH7O"

# ─────────────────────────────────────────────────────
#  Parsing helpers
# ─────────────────────────────────────────────────────

def parse_campaign_name(campaign_name: str):
    """From a campaign name return ('TC238', 'CA6.DIANA') or None."""
    if not campaign_name:
        return None

    tc_match = re.search(r'TC(\d+)', campaign_name, re.IGNORECASE)
    if not tc_match:
        return None
    tc_label = f"TC{tc_match.group(1)}"

    bracket_match = re.search(r'\]([^[]*?)TC\d+', campaign_name, re.IGNORECASE)
    if bracket_match:
        account_name = bracket_match.group(1).strip()
    else:
        account_name = ""

    return tc_label, account_name


def extract_ad_name_from_campaign(campaign_name: str) -> str:
    """Extracts the ad name from a campaign name. e.g. '... - LT801.30' -> 'LT801.30'"""
    if not campaign_name:
        return ""
    # Try the canonical ABO/CBO separator first
    match = re.search(r'(?:ABO|CBO)\s+\S+\s*-\s*(.+)$', campaign_name, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    # Fallback: everything after the last ' - '
    parts = campaign_name.rsplit(' - ', 1)
    if len(parts) == 2:
        return parts[1].strip()
    return ""


def build_col_a_label(campaign_name: str) -> str:
    """Return formatted 'TCxxx ACCOUNT' string."""
    result = parse_campaign_name(campaign_name)
    if not result:
        return ""
    tc_label, account_name = result
    return f"{tc_label} {account_name}".strip()


# ─────────────────────────────────────────────────────
#  Export logic
# ─────────────────────────────────────────────────────

def export_creative_tests_excel(
    account_id: str,
    date_start: str,
    date_end: str,
    fb_token: str = None,
    fb_api_instance=None,
    progress_callback=None,
):
    """
    Fetches campaigns, parses them, and generates an Excel file in memory.
    Returns: BytesIO object containing the Excel file, and the count of extracted ads.
    """
    if fb_api_instance:
        fb_api = fb_api_instance
    else:
        token = fb_token or DEFAULT_FB_TOKEN
        fb_api = FacebookAdsAPI(token)

    if progress_callback:
        progress_callback("Extraindo Campanhas Ativas do Facebook...")

    raw_data = fb_api.get_ad_insights(
        account_id,
        date_start,
        date_end,
        level='campaign',
        progress_callback=progress_callback,
    )

    if progress_callback:
        progress_callback(f"{len(raw_data)} campanhas processadas na API. Construindo planilha...")

    if not raw_data:
        raise RuntimeError("A API retornou 0 campanhas para esta conta e período selecionado.")

    # Process data and group by TC
    # Structure: { "TC238 CA6.DIANA": ["LT801.30", "LT801.31"], ... }
    tc_groups = {}
    
    for record in raw_data:
        campaign_name = record.get("campaign_name", "")
        if not campaign_name:
            continue
            
        group_label = build_col_a_label(campaign_name)
        if not group_label:
            continue # Not a TC campaign
            
        ad_name = extract_ad_name_from_campaign(campaign_name)
        if not ad_name:
            continue
            
        if group_label not in tc_groups:
            tc_groups[group_label] = []
        tc_groups[group_label].append(ad_name)

    # Sort groups so they appear in order (TC238, TC239, etc.)
    sorted_labels = sorted(tc_groups.keys(), key=lambda x: (x.replace("TC", "").split(" ")[0] if "TC" in x else x))

    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Testes Criativos"

    # Headers
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)
    
    ws.cell(row=1, column=1, value="Teste e Conta").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=2, value="Nome do Anúncio").fill = header_fill
    ws.cell(row=1, column=2).font = header_font

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40

    current_row = 2
    total_ads_exported = 0

    for label in sorted_labels:
        ads = tc_groups[label]
        start_row = current_row
        
        for ad in ads:
            # Write Name
            ws.cell(row=current_row, column=2, value=ad)
            
            # Write Label (we need to write it in the first cell of the range before merging)
            if current_row == start_row:
                cell = ws.cell(row=current_row, column=1, value=label)
            
            current_row += 1
            total_ads_exported += 1
            
        end_row = current_row - 1
        
        # Merge the TC label cells
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            
        # Format the merged (or single) cell
        merged_cell = ws.cell(row=start_row, column=1)
        merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        merged_cell.font = Font(bold=True)

    if total_ads_exported == 0:
        raise ValueError("Nenhum anúncio com nomenclatura de TC (ex: [LOTTOV7]... TC238 ...) foi encontrado nas campanhas exportadas.")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output, total_ads_exported
