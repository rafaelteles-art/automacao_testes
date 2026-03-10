import requests
import json
import openpyxl

EXCEL_FILE = r"C:\Preencher planilha\FB - LOTTO V7.xlsx"
MAIN_SHEET = "032026"
DATA_START_ROW = 4

api_key = "wB7qY69R0KVU9tl4TBaQ"
date_start = "2026-02-03" 
date_end = "2026-03-05" 

redtrack_map = {}
print(f"Fetching RedTrack Data for {date_start} to {date_end}...")
try:
    r_rt = requests.get('https://api.redtrack.io/report', params={
        'api_key': api_key,
        'date_from': date_start,
        'date_to': date_end,
        'group': 'rt_ad',
        'limit': 5000
    }, timeout=30)
    if r_rt.status_code == 200:
        for r_row in r_rt.json():
            rt_ad = str(r_row.get('rt_ad', '')).strip().lower()
            if rt_ad:
                redtrack_map[rt_ad] = float(r_row.get('convtype2', 0))
except Exception as e:
    print(e)
    
print(f"RedTrack Catalog has {len(redtrack_map)} items.")

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[MAIN_SHEET]

test_ads = ["lt1017.7", "lt1010.2", "lt899.22", "lt1033.2", "lt1011.4", "lt899.32", "lt899.33", "lt1034.4", "lt1069", "lt1070"]

print("\n--- Testing Exact Extraction Logic ---")
for row_idx in range(DATA_START_ROW, ws.max_row + 1):
    ad_name_cell = ws.cell(row=row_idx, column=2)
    ad_name_value = ad_name_cell.value

    if not ad_name_value or str(ad_name_value).strip() == "":
        continue

    search_term = str(ad_name_value).strip().lower()
    
    is_test_subject = any(ta in search_term for ta in test_ads)
    if not is_test_subject: continue

    vendas = 0
    match_type = "None"
    matched_key = ""
    
    # Exact replica of fill_creative_tests.py logic
    ad_name_lower = search_term
    if ad_name_lower in redtrack_map:
        vendas = redtrack_map[ad_name_lower]
        match_type = "Exact"
        matched_key = ad_name_lower
    else:
        first_part = ad_name_lower.split(" - ")[0].split(" ")[0]
        if first_part in redtrack_map:
            vendas = redtrack_map[first_part]
            match_type = "First Part Chunk"
            matched_key = first_part
        else:
            for rt_key, rt_val in sorted(redtrack_map.items(), key=lambda item: len(item[0]), reverse=True):
                if ad_name_lower.startswith(rt_key):
                    vendas = rt_val
                    match_type = "Longest Prefix"
                    matched_key = rt_key
                    break

    print(f"Planilha: '{ad_name_value}' -> Mapeou: '{matched_key}' ({match_type}) -> Vendas Resolvidas: {vendas}")
