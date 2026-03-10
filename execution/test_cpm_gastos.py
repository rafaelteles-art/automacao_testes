#!/usr/bin/env python3
"""
Test script to simulate filling Col A (TC), Col G (CPM), Col J (Gasto)
with USD -> BRL conversion, without saving the Excel file.
"""

import re
import sys
import os
import requests
from openpyxl import load_workbook

# App token and account
TOKEN = "EAAWDHozjODgBQZBEiBDPWYCCc5AmZA9FTkAh6ICQ3tMju4hCZAWCphJZCY9xsKZBIFq8PCefiiJnZA1MWVitYDNlIU86LOZAYkH91qPVx5vKuYNGb9xZCWGxcyAgp8qetdQgZB9ly3QyMz6wSyzck65iZCkWc8XAAlsAmG1nZB6z5quMXeDNfb5WMgE55N3yyGsUpyVsMEekMIYIl5nW0xH7jg2WUzT3ysBGEnZCrXVdQCTVc5iD6kqfuC5fATZAHqHXyfNiRENJRDHjulEgLszhvKegV7uKYfEBWBlrJeXZBtnGgZD"
ACCOUNT_ID = "act_542987171356461"  # CA06 Diana
EXCEL_FILE = r"C:\Preencher planilha\FB - LOTTO V7.xlsx"
MAIN_SHEET = "032026"
DATA_START_ROW = 4

# Date range for testing
DATE_START = "2026-03-01"
DATE_END = "2026-03-03"

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from fill_creative_tests import extract_ad_name_from_campaign, build_col_a_label
from facebook_redtrack_importer_v2 import FacebookAdsAPI

# 1. Fetch USD to BRL Quote
try:
    r_quote = requests.get("https://economia.awesomeapi.com.br/last/USD-BRL")
    usd_to_brl = float(r_quote.json()["USDBRL"]["ask"])
    print(f"✅ Cotação USD -> BRL obtida em tempo real: R$ {usd_to_brl:.4f}")
except Exception as e:
    print(f"⚠️ Erro ao obter cotação: {e}. Usando R$ 5.75 como fallback.")
    usd_to_brl = 5.75

fb_api = FacebookAdsAPI(TOKEN)

# 2. Fetch /campaigns to build naming catalog
print(f"Buscando Catálogo de Campanhas da conta {ACCOUNT_ID}...")
url = f"https://graph.facebook.com/v19.0/{ACCOUNT_ID}/campaigns"
all_campaigns = []
params = {"access_token": TOKEN, "fields": "id,name,effective_status", "limit": 500}
while url:
    r = requests.get(url, params=params, timeout=30)
    data = r.json()
    page = data.get("data", [])
    if not page: break
    all_campaigns.extend(page)
    url = data.get("paging", {}).get("next")
    params = {}

ad_to_campaign = {}
for camp in all_campaigns:
    name = camp.get("name", "")
    c_id = camp.get("id")
    if name:
        ext = extract_ad_name_from_campaign(name)
        if ext:
            key = ext.strip().lower()
            if key not in ad_to_campaign:
                ad_to_campaign[key] = {"name": name, "id": c_id}

# 3. Fetch /insights at campaign level to get Spend and CPM
# (Will only return campaigns that actually spent money)
print(f"Buscando Insights Financeiros ({DATE_START} até {DATE_END})...")
insights_data = fb_api.get_ad_insights(
    ACCOUNT_ID,
    DATE_START, # Need recent dates, but /campaigns will find all
    DATE_END,
    level='campaign'
)
finance_map: dict[str, dict] = {}  # Map campaign_id -> {spend, cpm}
for row in insights_data:
    c_id = row.get("campaign_id")
    if c_id:
        finance_map[c_id] = {
            "spend": float(row.get("spend", 0.0)),
            "cpm": float(row.get("cpm", 0.0))
        }

# 4. Read Excel and Simulate Output
print(f"Lendo Coluna B da planilha {MAIN_SHEET}...\n")
wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb[MAIN_SHEET]

print("=== RESULTADO DA SIMULAÇÃO ===")
print("LINHA | COL A (TC+Conta)       | COL B (Ad Paste)   | COL G (CPM R$) | COL J (Gasto R$)")
print("-" * 80)

for row_idx in range(DATA_START_ROW, ws.max_row + 1):
    ad_cell = ws.cell(row=row_idx, column=2).value
    if not ad_cell or str(ad_cell).strip() == "":
        continue

    search_term = str(ad_cell).strip().lower()
    matched_info = None

    if search_term in ad_to_campaign:
        matched_info = ad_to_campaign[search_term]
    else:
        for key, info in ad_to_campaign.items():
            if search_term in key or key in search_term:
                matched_info = info
                break
    
    col_a_val = "NÃO ENCONTRADO"
    col_g_val = "0.00"
    col_j_val = "0.00"

    if matched_info:
        c_name = matched_info["name"]
        c_id = matched_info["id"]
        
        # Parse Col A
        col_a_label = build_col_a_label(c_name)
        if col_a_label:
            col_a_val = col_a_label

        # Fetch Financials
        fin = finance_map.get(c_id, {"spend": 0.0, "cpm": 0.0})
        
        # Convert to BRL
        spend_brl = fin["spend"] * usd_to_brl
        cpm_brl = fin["cpm"] * usd_to_brl

        col_j_val = f"{spend_brl:.2f}"
        col_g_val = f"{cpm_brl:.2f}"

    print(f"L{row_idx:03d}  | {col_a_val:<22} | {str(ad_cell):<18} | R$ {col_g_val:>8} | R$ {col_j_val:>8}")

print("-" * 80)
print("Fim da simulação.")
