"""Quick scan of columns N-W to find where the new section data lives"""
from openpyxl import load_workbook

EXCEL_FILE = r"C:\Preencher planilha\FB - LOTTO V7.xlsx"
wb = load_workbook(EXCEL_FILE, read_only=True)
ws = wb["032026"]

# Print headers (rows 1-3)
print("=== Headers (rows 1-3) for columns N-W ===")
for row_idx in range(1, 4):
    vals = []
    for col in range(14, 23):  # N=14 to W=23
        cell = ws.cell(row=row_idx, column=col)
        val = str(cell.value)[:20] if cell.value else ""
        vals.append(f"{val:<20}")
    print(f"Row {row_idx}: {' | '.join(vals)}")

# Print first 30 data rows
print("\n=== Data rows 4-33 for cols N-W ===")
print(f"{'Row':<5} | {'N(14)':<15} | {'O(15)':<15} | {'P(16)':<15} | {'Q(17)':<15} | {'R(18)':<15} | {'S(19)':<15} | {'T(20)':<15} | {'U(21)':<15} | {'V(22)':<15}")
print("-" * 160)

for row_idx in range(4, 34):
    vals = []
    for col in range(14, 23):
        cell = ws.cell(row=row_idx, column=col)
        val = str(cell.value)[:15] if cell.value else ""
        vals.append(f"{val:<15}")
    # Only print if at least one column has data
    if any(ws.cell(row=row_idx, column=c).value for c in range(14, 23)):
        print(f"{row_idx:<5} | {' | '.join(vals)}")

wb.close()
