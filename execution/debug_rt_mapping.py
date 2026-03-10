import requests
import json
import openpyxl

EXCEL_FILE = r"C:\Preencher planilha\FB - LOTTO V7.xlsx"
MAIN_SHEET = "032026"
DATA_START_ROW = 4

api_key = "wB7qY69R0KVU9tl4TBaQ"

# Simulando a janela de 30 dias que o web_app usa por padrão
date_start = "2026-02-03" # datetime.now() - 30 days
date_end = "2026-03-05" 

redtrack_map = {}
print("Fetching RedTrack Data...")
try:
    r_rt = requests.get('https://api.redtrack.io/report', params={
        'api_key': api_key,
        'date_from': date_start,
        'date_to': date_end,
        'group': 'rt_ad',
        'limit': 5000
    }, timeout=30)
    if r_rt.status_code == 200:
        for r_row in r_rt.json():
            rt_ad = str(r_row.get('rt_ad', '')).strip().lower()
            if rt_ad:
                redtrack_map[rt_ad] = float(r_row.get('convtype2', 0))
except Exception as e:
    print(e)
    
print(f"RedTrack Catalog has {len(redtrack_map)} items.")

print("\n--- Testing Excel Substring Mapping ---")
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
ws = wb[MAIN_SHEET]

test_ads = ["lt1017.7", "lt1010.2", "lt899.22", "lt1033.2", "lt1011.4", "lt899.32", "lt899.33", "lt1034.4", "lt1069", "lt1070"]

for row_idx in range(DATA_START_ROW, ws.max_row + 1):
    ad_name_cell = ws.cell(row=row_idx, column=2)
    ad_name_value = ad_name_cell.value

    if not ad_name_value or str(ad_name_value).strip() == "":
        continue

    search_term = str(ad_name_value).strip().lower()
    
    # Check if this row is one of our test subjects
    is_test_subject = False
    for ta in test_ads:
        if ta in search_term:
            is_test_subject = True
            break
            
    if not is_test_subject: continue

    vendas = 0
    match_type = "None"
    matched_key = ""
    
    if search_term in redtrack_map:
        vendas = redtrack_map[search_term]
        match_type = "Exact"
        matched_key = search_term
    else:
        # Fallback partial match (the bug might be here!)
        for rt_key, rt_val in redtrack_map.items():
            if rt_key in search_term:
                vendas = rt_val
                match_type = "Partial substring"
                matched_key = rt_key
                break

    print(f"Row {row_idx} Col B: '{ad_name_value}' -> Map Matched: {matched_key} ({match_type}) -> Vendas: {vendas}")
