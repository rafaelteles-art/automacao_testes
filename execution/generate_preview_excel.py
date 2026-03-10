#!/usr/bin/env python3
"""
Generates a physical preview Excel file duplicating the user's conditional logic.
Reads FB - LOTTO V7.xlsx, updates only rows with "TESTE" in Col M, and saves as FB - LOTTO V7_Preview.xlsx
"""

import sys, os, requests
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from fill_creative_tests import extract_ad_name_from_campaign, build_col_a_label

TOKEN = "EAAWDHozjODgBQZBEiBDPWYCCc5AmZA9FTkAh6ICQ3tMju4hCZAWCphJZCY9xsKZBIFq8PCefiiJnZA1MWVitYDNlIU86LOZAYkH91qPVx5vKuYNGb9xZCWGxcyAgp8qetdQgZB9ly3QyMz6wSyzck65iZCkWc8XAAlsAmG1nZB6z5quMXeDNfb5WMgE55N3yyGsUpyVsMEekMIYIl5nW0xH7jg2WUzT3ysBGEnZCrXVdQCTVc5iD6kqfuC5fATZAHqHXyfNiRENJRDHjulEgLszhvKegV7uKYfEBWBlrJeXZBtnGgZD"
ACCOUNT_ID = "act_542987171356461"
ORIGINAL_FILE = r"C:\Preencher planilha\FB - LOTTO V7.xlsx"
PREVIEW_FILE = r"C:\Preencher planilha\FB - LOTTO V7_Preview.xlsx"
MAIN_SHEET = "032026"
DATA_START_ROW = 4

DATE_START = "2026-03-01"
DATE_END = "2026-03-03"
USD_TO_BRL = 5.2797

print("Buscando Catálogo de Campanhas...")
url = f"https://graph.facebook.com/v19.0/{ACCOUNT_ID}/campaigns"
all_campaigns = []
params = {"access_token": TOKEN, "fields": "id,name,effective_status", "limit": 500}
while url:
    r = requests.get(url, params=params)
    data = r.json()
    page = data.get("data", [])
    if not page: break
    all_campaigns.extend(page)
    url = data.get("paging", {}).get("next")
    params = {}

ad_to_campaign = {}
for camp in all_campaigns:
    name = camp.get("name", "")
    c_id = camp.get("id")
    if name:
        ext = extract_ad_name_from_campaign(name)
        if ext:
            key = ext.strip().lower()
            if key not in ad_to_campaign:
                ad_to_campaign[key] = {"id": c_id, "name": name}

print("Buscando Insights Financeiros...")
insights_url = f"https://graph.facebook.com/v19.0/{ACCOUNT_ID}/insights"
params = {
    'access_token': TOKEN,
    'fields': 'campaign_id,impressions,cpc,cpm,ctr,spend,actions,video_p75_watched_actions',
    'level': 'campaign',
    'date_start': DATE_START,
    'date_end': DATE_END,
    'limit': 100
}
finance_map = {}
while insights_url:
    r = requests.get(insights_url, params=params)
    data = r.json()
    page = data.get('data', [])
    if not page: break
    for row in page:
        c_id = row.get("campaign_id")
        imps = float(row.get("impressions", 0))
        
        actions = row.get('actions', [])
        video_3s_views = 0
        for act in actions:
            if act.get('action_type') == 'video_view':
                video_3s_views = float(act.get('value', 0))
                break
                
        video_p75_actions = row.get('video_p75_watched_actions', [])
        p75 = 0
        if video_p75_actions:
            p75 = float(video_p75_actions[0].get('value', 0))
            
        hook_rate = (video_3s_views / imps) if imps > 0 else 0
        body_rate = (p75 / imps) if imps > 0 else 0
        
        finance_map[c_id] = {
            "spend": float(row.get("spend", 0.0)),
            "cpm": float(row.get("cpm", 0.0)),
            "cpc": float(row.get("cpc", 0.0)),
            "ctr": (float(row.get("ctr", 0.0)) / 100), # Facebook returns CTR as a percentage string "3.14", so divide by 100 for Excel percentage
            "hook_rate": hook_rate,
            "body_rate": body_rate
        }
        
    insights_url = data.get("paging", {}).get("next")
    params = {}

print("Criando Cópia da Planilha...")
shutil.copy2(ORIGINAL_FILE, PREVIEW_FILE)

print("Preenchendo Cópia da Planilha...")
wb = load_workbook(PREVIEW_FILE)
ws = wb[MAIN_SHEET]

print("Desmesclando Coluna A...")
for merged_range in list(ws.merged_cells.ranges):
    if merged_range.min_col == 1 and merged_range.max_col == 1:
        ws.unmerge_cells(str(merged_range))

updated_count = 0
skipped_count = 0

for row_idx in range(DATA_START_ROW, ws.max_row + 1):
    ad_cell = ws.cell(row=row_idx, column=2).value # Col B
    if not ad_cell or str(ad_cell).strip() == "":
        continue # skip blank rows, don't stop entirely
        
    status_cell = ws.cell(row=row_idx, column=13).value # Col M
    current_status = str(status_cell).strip().upper() if status_cell else "(VAZIO)"
    
    if "TESTE" in current_status:
        # User requested update
        search_term = str(ad_cell).strip().lower()
        matched_info = None

        if search_term in ad_to_campaign:
            matched_info = ad_to_campaign[search_term]
        else:
            for key, info in ad_to_campaign.items():
                if search_term in key or key in search_term:
                    matched_info = info
                    break
                    
        if matched_info:
            c_id = matched_info["id"]
            c_name = matched_info["name"]
            
            # Fill Col A (TC + Account)
            ws.cell(row=row_idx, column=1, value=build_col_a_label(c_name))
            
            # Get Finances
            fin = finance_map.get(c_id, {
                "spend": 0.0, "cpm": 0.0, "cpc": 0.0,
                "ctr": 0.0, "hook_rate": 0.0, "body_rate": 0.0
            })
            
            cpc_brl = fin["cpc"] * USD_TO_BRL
            cpm_brl = fin["cpm"] * USD_TO_BRL
            spend_brl = fin["spend"] * USD_TO_BRL
            
            # Hook Rate (Col E = 5)
            c5 = ws.cell(row=row_idx, column=5, value=fin["hook_rate"])
            c5.number_format = '0.00%'
            
            # Body View 75% (Col F = 6)
            c6 = ws.cell(row=row_idx, column=6, value=fin["body_rate"])
            c6.number_format = '0.00%'
            
            # CPM (Col G = 7)
            c7 = ws.cell(row=row_idx, column=7, value=cpm_brl)
            c7.number_format = 'R$ #,##0.00'
            
            # CTR (Col H = 8)
            c8 = ws.cell(row=row_idx, column=8, value=fin["ctr"])
            c8.number_format = '0.00%'
            
            # CPC (Col I = 9)
            c9 = ws.cell(row=row_idx, column=9, value=cpc_brl)
            c9.number_format = 'R$ #,##0.00'
            
            # GASTO (Col J = 10)
            c10 = ws.cell(row=row_idx, column=10, value=spend_brl)
            c10.number_format = 'R$ #,##0.00'
            
            updated_count += 1
    else:
        skipped_count += 1

print("Remesclando Coluna A...")
row = DATA_START_ROW
while row <= ws.max_row:
    cell_val = ws.cell(row=row, column=1).value
    if not cell_val:
        row += 1
        continue

    end_row = row
    while end_row + 1 <= ws.max_row and ws.cell(end_row + 1, 1).value == cell_val:
        end_row += 1

    if end_row > row:
        ws.merge_cells(start_row=row, start_column=1, end_row=end_row, end_column=1)
        merged_cell = ws.cell(row=row, column=1)
        merged_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        merged_cell.font = Font(bold=True)

    row = end_row + 1

wb.save(PREVIEW_FILE)
print(f"\\nSUCESSO!")
print(f"Linhas atualizadas (TESTE): {updated_count}")
print(f"Linhas puladas    (OUTROS): {skipped_count}")
print(f"Arquivo salvo como: {PREVIEW_FILE}")
