#!/usr/bin/env python3
"""
Facebook Ads & RedTrack Data Importer - Version 2
Advanced version with interactive account selection and automated data mapping
"""

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json
import sys
from typing import List, Dict, Tuple

if sys.version_info >= (3, 7) and sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

class FacebookAdsAPI:
    """Facebook Ads API client"""
    
    def __init__(self, access_token: str):
        self.access_token = access_token
        self.base_url = "https://graph.facebook.com/v19.0"
        self.session = requests.Session()
        self.session.headers.update({'Authorization': f'Bearer {access_token}'})
    
    def get_business_managers(self) -> List[Dict]:
        """Get all business managers accessible to this token"""
        try:
            url = f"{self.base_url}/me/businesses"
            params = {'fields': 'id,name'}
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json().get('data', [])
        except Exception as e:
            print(f"❌ Error fetching business managers: {e}")
            return []
    
    def get_ad_accounts(self, business_manager_id: str) -> List[Dict]:
        """Get all ad accounts from a business manager"""
        try:
            url = f"{self.base_url}/{business_manager_id}/owned_ad_accounts"
            params = {
                'fields': 'id,name,currency,account_status',
                'limit': 100
            }
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json().get('data', [])
        except Exception as e:
            print(f"❌ Error fetching ad accounts: {e}")
            return []
    
    def get_ad_insights(self, account_id: str, date_start: str, date_end: str, level: str = 'ad', progress_callback=None) -> List[Dict]:
        """Get detailed ad insights for a date range"""
        # Clean account ID
        if account_id.startswith('act_'):
            account_id = account_id[4:]
            
        fields_str = 'campaign_id,campaign_name,impressions,clicks,spend,cpc,cpm,ctr'
        if level == 'ad':
            fields_str = 'campaign_id,campaign_name,ad_id,ad_name,impressions,clicks,spend,cpc,cpm,ctr'
        
        url = f"{self.base_url}/act_{account_id}/insights"
        params = {
            'access_token': self.access_token,
            'fields': fields_str,
            'level': level,
            'time_range': f'{{"since":"{date_start}","until":"{date_end}"}}',
            'limit': 100
        }
        
        all_data = []
        try:
            while url:
                response = self.session.get(url, params=params, timeout=30)
                response.raise_for_status()
                data = response.json()
                
                page_data = data.get('data', [])
                if not page_data:
                    break # Safety mechanism against FB API infinite empty cursor bug
                    
                all_data.extend(page_data)
                
                # Feedback to Streamlit UI
                if progress_callback:
                    progress_callback(len(all_data))
                
                # Check for next page URL
                paging = data.get('paging', {})
                url = paging.get('next')
                params = {} # The 'next' URL already contains all parameters
            
            return all_data
        except Exception as e:
            print(f"❌ Error fetching ad insights: {e}")
            if progress_callback:
                progress_callback(f"⚠️ Erro de conexão após baixar {len(all_data)} linhas. Exibindo dados incompletos.")
            return all_data


class RedTrackAPI:
    """RedTrack API client"""
    
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = "https://api.redtrack.io"
        self.session = requests.Session()
        self.session.headers.update({
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        })
    
    def get_conversions(self, date_start: str, date_end: str) -> List[Dict]:
        """Get conversion data"""
        try:
            url = f"{self.base_url}/conversions/export"
            params = {
                'date_from': date_start,
                'date_to': date_end,
                'limit': 10000
            }
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json().get('data', [])
        except Exception as e:
            print(f"❌ Error fetching conversions: {e}")
            return []
    
    def get_campaigns(self, date_start: str, date_end: str) -> List[Dict]:
        """Get campaign statistics"""
        try:
            url = f"{self.base_url}/campaigns"
            params = {
                'date_from': date_start,
                'date_to': date_end,
                'limit': 1000
            }
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json().get('data', [])
        except Exception as e:
            print(f"❌ Error fetching campaigns: {e}")
            return []


class ExcelManager:
    """Excel file management"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.raw_sheet = None
    
    def create_raw_data_sheet(self) -> None:
        """Create raw data sheet"""
        sheet_name = "Dados Brutos"
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]
        self.raw_sheet = self.wb.create_sheet(sheet_name, 0)
    
    def add_facebook_data(self, data: List[Dict]) -> int:
        """Add Facebook data to raw sheet. Returns number of rows added."""
        if not data:
            return 0
        
        df = pd.DataFrame(data)
        
        # Add headers
        for col_idx, col_name in enumerate(df.columns, 1):
            self.raw_sheet.cell(row=1, column=col_idx, value=col_name)
        
        # Add data rows
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                self.raw_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        return len(df)
    
    def add_redtrack_data(self, data: List[Dict], start_column: int) -> int:
        """Add RedTrack data to raw sheet. Returns number of rows added."""
        if not data:
            return 0
        
        df = pd.DataFrame(data)
        
        # Add headers
        for col_idx, col_name in enumerate(df.columns, start_column):
            self.raw_sheet.cell(row=1, column=col_idx, value=col_name)
        
        # Add data rows
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, start_column):
                self.raw_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        return len(df)
    
    def save(self) -> None:
        """Save workbook"""
        self.wb.save(self.file_path)
        print(f"✓ Workbook saved: {self.file_path}")


def select_account_interactively(fb_api: FacebookAdsAPI) -> Tuple[str, str]:
    """Interactive account selection"""
    
    print("\n" + "="*60)
    print("ACCOUNT SELECTION")
    print("="*60)
    
    # Get business managers
    bms = fb_api.get_business_managers()
    if not bms:
        print("❌ No business managers found")
        return None, None
    
    print(f"\nFound {len(bms)} Business Manager(s):")
    for idx, bm in enumerate(bms, 1):
        print(f"{idx}. {bm['name']} (ID: {bm['id']})")
    
    # Select BM
    bm_choice = input(f"\nSelect Business Manager (1-{len(bms)}): ").strip()
    try:
        bm_idx = int(bm_choice) - 1
        if 0 <= bm_idx < len(bms):
            selected_bm = bms[bm_idx]
        else:
            print("❌ Invalid selection")
            return None, None
    except ValueError:
        print("❌ Invalid input")
        return None, None
    
    # Get ad accounts
    accounts = fb_api.get_ad_accounts(selected_bm['id'])
    if not accounts:
        print("❌ No ad accounts found in this BM")
        return None, None
    
    print(f"\nFound {len(accounts)} Ad Account(s):")
    for idx, acc in enumerate(accounts, 1):
        status = acc.get('account_status', 'unknown')
        print(f"{idx}. {acc['name']} (ID: {acc['id']}) - Status: {status}")
    
    # Select account
    acc_choice = input(f"\nSelect Ad Account (1-{len(accounts)}): ").strip()
    try:
        acc_idx = int(acc_choice) - 1
        if 0 <= acc_idx < len(accounts):
            selected_acc = accounts[acc_idx]
        else:
            print("❌ Invalid selection")
            return None, None
    except ValueError:
        print("❌ Invalid input")
        return None, None
    
    return selected_bm['id'], selected_acc['id']


def main(account_id=None, date_start=None, date_end=None, fb_token=None, rt_token=None):
    """Main execution"""
    
    # Configuration
    FB_ACCESS_TOKEN = fb_token or os.environ.get("FB_ACCESS_TOKEN", "EAAWDHozjODgBQ0b4ZAZBOZBzGhqi9ZCX0bj8DbmAPnsBfYbEMMZCqMeBMCmLjB2dpzxHvzZC6UQGApi9frZAWyQHPmHZB1hFJa2q3nTNaaDtwHSxqJB5Veeo1CpE9gTYAD3vpJf9vRNNj62z2ebVJ6tD0mKbIzh9DXZCbrnjOHhiAkrcffsEwKcZAuHchAMZBRgi1BjmUIjP2IhfH7O")
    REDTRACK_API_KEY = rt_token or "wB7qY69R0KVU9tl4TBaQ"
    EXCEL_FILE = r"C:\Preencher planilha\FB - LOTTO V7.xlsx"
    
    print("\n" + "="*60)
    print("Facebook Ads & RedTrack Data Importer v2")
    print("="*60)
    
    # Initialize APIs
    fb_api = FacebookAdsAPI(FB_ACCESS_TOKEN)
    rt_api = RedTrackAPI(REDTRACK_API_KEY)
    excel_mgr = ExcelManager(EXCEL_FILE)
    
    # Select account
    if not account_id:
        bm_id, account_id = select_account_interactively(fb_api)
        if not account_id:
            print("❌ Account selection failed")
            return
    
    print(f"\n✓ Selected Account: {account_id}")
    
    # Date range
    if not date_end:
        date_end = datetime.now().strftime('%Y-%m-%d')
    if not date_start:
        date_start = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    print(f"✓ Date Range: {date_start} to {date_end}")
    
    # Create raw data sheet
    print("\n1. Creating raw data sheet...")
    excel_mgr.create_raw_data_sheet()
    
    # Fetch Facebook data
    print(f"2. Fetching Facebook Ads data...")
    fb_data = fb_api.get_ad_insights(account_id, date_start, date_end)
    fb_rows = excel_mgr.add_facebook_data(fb_data)
    print(f"   ✓ Added {fb_rows} rows of Facebook data")
    
    # Fetch RedTrack data
    print(f"3. Fetching RedTrack conversion data...")
    rt_data = rt_api.get_conversions(date_start, date_end)
    rt_rows = excel_mgr.add_redtrack_data(rt_data, start_column=len(fb_data[0]) + 2 if fb_data else 2)
    print(f"   ✓ Added {rt_rows} rows of RedTrack data")
    
    # Save
    print(f"\n4. Saving Excel file...")
    excel_mgr.save()
    
    print("\n" + "="*60)
    print("✓ Import completed successfully!")
    print("="*60)
    print(f"\nSummary:")
    print(f"  • Facebook Ads records: {fb_rows}")
    print(f"  • RedTrack conversions: {rt_rows}")
    print(f"  • File: {EXCEL_FILE}")


    return fb_data, rt_data


if __name__ == "__main__":
    main()
