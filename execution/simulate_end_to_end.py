import sys
import os
import shutil
import openpyxl

# Overwrite EXCEL_FILE to a temporary simulated file
import fill_creative_tests
ORIGINAL_EXCEL = r"C:\Preencher planilha\FB - LOTTO V7.xlsx"
SIMULATED_EXCEL = r"C:\Preencher planilha\FB - LOTTO V7_SIMULADO.xlsx"
fill_creative_tests.EXCEL_FILE = SIMULATED_EXCEL

print("Criando cópia da planilha original para simulação segura...")
shutil.copyfile(ORIGINAL_EXCEL, SIMULATED_EXCEL)

def mock_progress(msg):
    print(f"[STATUS] {msg}")

from facebook_redtrack_importer_v2 import FacebookAdsAPI
fb_token = "EAAWDHozjODgBQwIYZAMRFBq5IyKIqff4ngn2S7Ca7j6apNVGvf7kdNnt1W6wJA8XhO1G3p5hA0SJOkPZABtcdKqPHSdhxvnZBruHuRZBb6gX3ZB1gVrsUp19cDA5cVjGhy3FZBZCjJjNVAiZBFZA55ZBZBk2BpSbBiVPngJgJoomdSRe8o33CstSyK8yZCX6c53NZBhlzZCcJWlGdj2H72TkvFYgl4NkZAS2fnM3PFgkPAHVIUu3GIfDx99wq9EfZCKEKqOlHI7PcLXhsfdCZClmOZAjHUWcIgHTocwvUZCKIuWK3vfCO4ZD"
fb = FacebookAdsAPI(fb_token)

acc_ids = [
    "act_987248712293933",
    "act_1583963688753767",
    "act_6481076258591934",
    "act_1230367020974448",
    "act_542987171356461",
    "act_1300700817490669",
    "act_854290605799299",
    "act_982787082695569"
]

print(f"Executando preenchimento com as contas: {acc_ids}")

try:
    res = fill_creative_tests.fill_creative_tests(
        account_ids=acc_ids,
        date_start="2024-01-01",  # Puxando um histórico gigante para garantir captura de conversões
        date_end="2026-04-01",
        fb_token=fb_token,
        redtrack_token="wB7qY69R0KVU9tl4TBaQ",
        progress_callback=mock_progress
    )
    print("\nPreenchimento no arquivo SIMULADO finalizado!")
    
except Exception as e:
    import traceback
    print("ERRO OCORRIDO DURANTE PREENCHIMENTO:")
    traceback.print_exc()

# Agora vamos ler a SIMULAÇÃO e imprimir as linhas alvo
print("\n" + "="*70)
print("📊 RESULTADO DA SIMULAÇÃO (Extraído direto do FB - LOTTO V7_SIMULADO.xlsx)")
print("="*70)

try:
    wb = openpyxl.load_workbook(SIMULATED_EXCEL, data_only=True)
    ws = wb['032026']
    test_ads = ['lt1017.7', 'lt1010.2', 'lt899.22', 'lt1033.2', 'lt1011.4', 'lt899.32', 'lt899.33', 'lt1034.4', 'lt1069', 'lt1070']
    
    for r in range(4, ws.max_row + 1):
        ad = ws.cell(row=r, column=2).value
        status = ws.cell(row=r, column=13).value
        
        if not ad: continue
        ad_str = str(ad).strip()
        
        if any(t in ad_str.lower() for t in test_ads):
            vendas = ws.cell(row=r, column=11).value
            cpa = ws.cell(row=r, column=12).value
            gasto = ws.cell(row=r, column=10).value
            
            vendas_str = str(vendas) if vendas is not None else "N/A"
            
            # Se a linha não tinha "TESTE" na Coluna M, o Vendas fica N/A pro robô.
            # O Excel base pode já ter algo, ou ser branco.
            status_str = str(status) if status else "Vazio"
            
            if isinstance(cpa, (int, float)):
                cpa_str = f"R$ {cpa:.2f}"
            else:
                cpa_str = str(cpa)
                
            if isinstance(gasto, (int, float)):
                gasto_str = f"R$ {gasto:.2f}"
            else:
                gasto_str = str(gasto)
                
            print(f"Row {r:4d} | [{ad_str:<20}] | Status (Col M): {status_str:<10} | Gasto: {gasto_str:<12} | Vendas (K): {vendas_str:<5} | CPA (L): {cpa_str}")
except Exception as e:
    print("Erro lendo simulação:", e)

# Apagar a simulação
if os.path.exists(SIMULATED_EXCEL):
    os.remove(SIMULATED_EXCEL)
